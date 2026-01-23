from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from jose import JWTError, jwt
from passlib.context import CryptContext
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from io import BytesIO
import pandas as pd

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        role: str = payload.get("role")
        employee_id: str = payload.get("employee_id")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        return {"username": username, "role": role, "employee_id": employee_id}
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    role: str
    employee_id: Optional[str] = None

class Employee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    employee_id: str
    full_name: str
    department: str
    designation: str
    email: EmailStr
    date_of_joining: str
    status: str

class EmployeeCreate(BaseModel):
    full_name: str
    department: str
    designation: str
    email: EmailStr
    date_of_joining: str
    status: str = "Active"

class Asset(BaseModel):
    model_config = ConfigDict(extra="ignore")
    asset_id: str
    asset_name: str
    category: str
    brand: str
    serial_number: Optional[str] = None
    imei_2: Optional[str] = None
    condition: str
    status: str

class AssetCreate(BaseModel):
    asset_name: str
    category: str
    brand: str
    serial_number: Optional[str] = None
    imei_2: Optional[str] = None
    condition: str = "New"
    status: str = "Available"

class Assignment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    assignment_id: str
    employee_id: str
    employee_name: str
    asset_id: str
    asset_name: str
    assigned_date: str
    return_date: Optional[str] = None
    remarks: Optional[str] = None
    sim_provider: Optional[str] = None
    sim_mobile_number: Optional[str] = None
    sim_type: Optional[str] = None
    sim_ownership: Optional[str] = None
    sim_purpose: Optional[str] = None

class AssignmentCreate(BaseModel):
    employee_id: str
    asset_id: str
    assigned_date: str
    return_date: Optional[str] = None
    remarks: Optional[str] = None
    sim_provider: Optional[str] = None
    sim_mobile_number: Optional[str] = None
    sim_type: Optional[str] = None
    sim_ownership: Optional[str] = None
    sim_purpose: Optional[str] = None

class DashboardStats(BaseModel):
    total_assets: int
    assigned_assets: int
    available_assets: int
    total_employees: int

class PendingReturn(BaseModel):
    employee_id: str
    employee_name: str
    email: str
    assets: List[dict]

@api_router.post("/auth/login", response_model=LoginResponse)
async def login(login_req: LoginRequest):
    user = await db.users.find_one({"username": login_req.username}, {"_id": 0})
    if not user or not verify_password(login_req.password, user["password"]):
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    
    token_data = {
        "sub": user["username"],
        "role": user.get("role", "HR"),
        "employee_id": user.get("employee_id")
    }
    access_token = create_access_token(data=token_data)
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.get("role", "HR"),
        "employee_id": user.get("employee_id")
    }

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

@api_router.get("/global-search")
async def global_search(q: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    results = {"employees": [], "assets": []}
    
    # Search employees
    employee_query = {
        "$or": [
            {"full_name": {"$regex": q, "$options": "i"}},
            {"employee_id": {"$regex": q, "$options": "i"}}
        ]
    }
    employees = await db.employees.find(employee_query, {"_id": 0}).to_list(10)
    
    for employee in employees:
        assignments = await db.assignments.find(
            {"employee_id": employee["employee_id"], "return_date": None},
            {"_id": 0}
        ).to_list(100)
        employee["assigned_assets"] = assignments
    
    results["employees"] = employees
    
    # Search assets
    asset_query = {
        "$or": [
            {"asset_name": {"$regex": q, "$options": "i"}},
            {"category": {"$regex": q, "$options": "i"}},
            {"serial_number": {"$regex": q, "$options": "i"}},
            {"imei_2": {"$regex": q, "$options": "i"}},
            {"asset_id": {"$regex": q, "$options": "i"}}
        ]
    }
    assets = await db.assets.find(asset_query, {"_id": 0}).to_list(10)
    
    for asset in assets:
        if asset["status"] == "Assigned":
            assignment = await db.assignments.find_one(
                {"asset_id": asset["asset_id"], "return_date": None},
                {"_id": 0}
            )
            asset["assigned_to"] = assignment
        else:
            asset["assigned_to"] = None
    
    results["assets"] = assets
    
    return results

@api_router.get("/employees", response_model=List[Employee])
async def get_employees(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    return employees

@api_router.get("/employees/me", response_model=Employee)
async def get_my_profile(current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "Employee" and current_user["employee_id"]:
        employee = await db.employees.find_one({"employee_id": current_user["employee_id"]}, {"_id": 0})
        if not employee:
            raise HTTPException(status_code=404, detail="Employee profile not found")
        return employee
    raise HTTPException(status_code=403, detail="Access denied")

@api_router.post("/employees", response_model=Employee)
async def create_employee(employee: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    count = await db.employees.count_documents({})
    employee_id = f"EMP{str(count + 1).zfill(4)}"
    
    employee_dict = employee.model_dump()
    employee_dict["employee_id"] = employee_id
    
    await db.employees.insert_one(employee_dict)
    return Employee(**employee_dict)

@api_router.post("/employees/import")
async def import_employees(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        required_columns = ['Full Name', 'Department', 'Designation', 'Email', 'Date of Joining', 'Status']
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(status_code=400, detail=f"Excel file must contain columns: {', '.join(required_columns)}")
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                count = await db.employees.count_documents({})
                employee_id = f"EMP{str(count + 1).zfill(4)}"
                
                employee_data = {
                    "employee_id": employee_id,
                    "full_name": str(row['Full Name']),
                    "department": str(row['Department']),
                    "designation": str(row['Designation']),
                    "email": str(row['Email']),
                    "date_of_joining": str(row['Date of Joining']),
                    "status": str(row['Status'])
                }
                
                await db.employees.insert_one(employee_data)
                imported_count += 1
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return {
            "message": f"Successfully imported {imported_count} employees",
            "imported": imported_count,
            "errors": errors
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.get("/employees/export")
async def export_employees(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    headers = ["Employee ID", "Full Name", "Department", "Designation", "Email", "Date of Joining", "Status"]
    ws.append(headers)
    
    for employee in employees:
        ws.append([
            employee.get("employee_id", ""),
            employee.get("full_name", ""),
            employee.get("department", ""),
            employee.get("designation", ""),
            employee.get("email", ""),
            employee.get("date_of_joining", ""),
            employee.get("status", "")
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees.xlsx"}
    )

@api_router.get("/employees/template")
async def download_employees_template(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees Template"
    
    headers = ["Full Name", "Department", "Designation", "Email", "Date of Joining", "Status"]
    ws.append(headers)
    
    # Add sample data
    ws.append(["John Smith", "IT", "Software Engineer", "john.smith@example.com", "2024-01-15", "Active"])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees_template.xlsx"}
    )

@api_router.put("/employees/{employee_id}", response_model=Employee)
async def update_employee(employee_id: str, employee: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    employee_dict = employee.model_dump()
    result = await db.employees.update_one({"employee_id": employee_id}, {"$set": employee_dict})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    employee_dict["employee_id"] = employee_id
    return Employee(**employee_dict)

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = await db.employees.delete_one({"employee_id": employee_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee deleted successfully"}

@api_router.get("/assets", response_model=List[Asset])
async def get_assets(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    assets = await db.assets.find({}, {"_id": 0}).to_list(1000)
    return assets

@api_router.post("/assets", response_model=Asset)
async def create_asset(asset: AssetCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    count = await db.assets.count_documents({})
    asset_id = f"AST{str(count + 1).zfill(4)}"
    
    asset_dict = asset.model_dump()
    asset_dict["asset_id"] = asset_id
    
    await db.assets.insert_one(asset_dict)
    return Asset(**asset_dict)

@api_router.post("/assets/import")
async def import_assets(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        required_columns = ['Asset Name', 'Category', 'Brand', 'Serial Number', 'Condition', 'Status']
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(status_code=400, detail=f"Excel file must contain columns: {', '.join(required_columns)}")
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                count = await db.assets.count_documents({})
                asset_id = f"AST{str(count + 1).zfill(4)}"
                
                asset_data = {
                    "asset_id": asset_id,
                    "asset_name": str(row['Asset Name']),
                    "category": str(row['Category']),
                    "brand": str(row['Brand']),
                    "serial_number": str(row['Serial Number']),
                    "condition": str(row['Condition']),
                    "status": str(row['Status'])
                }
                
                # Add IMEI 2 if present
                if 'IMEI 2' in df.columns and pd.notna(row.get('IMEI 2')):
                    asset_data["imei_2"] = str(row['IMEI 2'])
                
                await db.assets.insert_one(asset_data)
                imported_count += 1
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return {
            "message": f"Successfully imported {imported_count} assets",
            "imported": imported_count,
            "errors": errors
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.get("/assets/export")
async def export_assets(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    assets = await db.assets.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    
    headers = ["Asset ID", "Asset Name", "Category", "Brand", "Serial Number / IMEI 1", "IMEI 2", "Condition", "Status"]
    ws.append(headers)
    
    for asset in assets:
        ws.append([
            asset.get("asset_id", ""),
            asset.get("asset_name", ""),
            asset.get("category", ""),
            asset.get("brand", ""),
            asset.get("serial_number", ""),
            asset.get("imei_2", ""),
            asset.get("condition", ""),
            asset.get("status", "")
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assets.xlsx"}
    )

@api_router.get("/assets/template")
async def download_assets_template(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets Template"
    
    headers = ["Asset Name", "Category", "Brand", "Serial Number", "IMEI 2", "Condition", "Status"]
    ws.append(headers)
    
    # Add sample data
    ws.append(["Dell Laptop", "Electronics", "Dell", "DL123456", "", "New", "Available"])
    ws.append(["iPhone 15", "Mobile", "Apple", "356789012345678", "356789012345679", "New", "Available"])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assets_template.xlsx"}
    )

@api_router.put("/assets/{asset_id}", response_model=Asset)
async def update_asset(asset_id: str, asset: AssetCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    asset_dict = asset.model_dump()
    result = await db.assets.update_one({"asset_id": asset_id}, {"$set": asset_dict})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    asset_dict["asset_id"] = asset_id
    return Asset(**asset_dict)

@api_router.delete("/assets/{asset_id}")
async def delete_asset(asset_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = await db.assets.delete_one({"asset_id": asset_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Asset not found")
    return {"message": "Asset deleted successfully"}

@api_router.get("/assignments", response_model=List[Assignment])
async def get_assignments(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    assignments = await db.assignments.find({}, {"_id": 0}).to_list(1000)
    return assignments

@api_router.get("/assignments/my", response_model=List[Assignment])
async def get_my_assignments(current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "Employee" and current_user["employee_id"]:
        assignments = await db.assignments.find(
            {"employee_id": current_user["employee_id"]},
            {"_id": 0}
        ).to_list(1000)
        return assignments
    raise HTTPException(status_code=403, detail="Access denied")

@api_router.post("/assignments", response_model=Assignment)
async def create_assignment(assignment: AssignmentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    employee = await db.employees.find_one({"employee_id": assignment.employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    asset = await db.assets.find_one({"asset_id": assignment.asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    if asset["status"] == "Assigned":
        raise HTTPException(status_code=400, detail="Asset is already assigned")
    
    count = await db.assignments.count_documents({})
    assignment_id = f"ASG{str(count + 1).zfill(4)}"
    
    assignment_dict = assignment.model_dump()
    assignment_dict["assignment_id"] = assignment_id
    assignment_dict["employee_name"] = employee["full_name"]
    assignment_dict["asset_name"] = asset["asset_name"]
    
    await db.assignments.insert_one(assignment_dict)
    await db.assets.update_one({"asset_id": assignment.asset_id}, {"$set": {"status": "Assigned"}})
    
    return Assignment(**assignment_dict)

@api_router.put("/assignments/{assignment_id}", response_model=Assignment)
async def update_assignment(assignment_id: str, assignment: AssignmentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    existing = await db.assignments.find_one({"assignment_id": assignment_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    employee = await db.employees.find_one({"employee_id": assignment.employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    asset = await db.assets.find_one({"asset_id": assignment.asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    assignment_dict = assignment.model_dump()
    assignment_dict["employee_name"] = employee["full_name"]
    assignment_dict["asset_name"] = asset["asset_name"]
    
    if assignment.return_date and not existing.get("return_date"):
        await db.assets.update_one({"asset_id": assignment.asset_id}, {"$set": {"status": "Available"}})
    elif not assignment.return_date and existing.get("return_date"):
        await db.assets.update_one({"asset_id": assignment.asset_id}, {"$set": {"status": "Assigned"}})
    
    await db.assignments.update_one({"assignment_id": assignment_id}, {"$set": assignment_dict})
    
    assignment_dict["assignment_id"] = assignment_id
    return Assignment(**assignment_dict)

@api_router.delete("/assignments/{assignment_id}")
async def delete_assignment(assignment_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    assignment = await db.assignments.find_one({"assignment_id": assignment_id}, {"_id": 0})
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    await db.assets.update_one({"asset_id": assignment["asset_id"]}, {"$set": {"status": "Available"}})
    await db.assignments.delete_one({"assignment_id": assignment_id})
    
    return {"message": "Assignment deleted successfully"}

@api_router.post("/assignments/import")
async def import_assignments(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        # Check for required columns (flexible - either ID or Email/Serial)
        has_employee_id = 'Employee ID' in df.columns
        has_employee_email = 'Employee Email' in df.columns
        has_asset_id = 'Asset ID' in df.columns
        has_serial_number = 'Asset Serial Number' in df.columns
        
        if not (has_employee_id or has_employee_email):
            raise HTTPException(status_code=400, detail="Excel file must contain either 'Employee ID' or 'Employee Email' column")
        
        if not (has_asset_id or has_serial_number):
            raise HTTPException(status_code=400, detail="Excel file must contain either 'Asset ID' or 'Asset Serial Number' column")
        
        if 'Assigned Date' not in df.columns:
            raise HTTPException(status_code=400, detail="Excel file must contain 'Assigned Date' column")
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Find employee
                employee = None
                if has_employee_id and pd.notna(row.get('Employee ID')):
                    employee = await db.employees.find_one({"employee_id": str(row['Employee ID'])}, {"_id": 0})
                elif has_employee_email and pd.notna(row.get('Employee Email')):
                    employee = await db.employees.find_one({"email": str(row['Employee Email'])}, {"_id": 0})
                
                if not employee:
                    errors.append(f"Row {index + 2}: Employee not found")
                    continue
                
                # Find asset
                asset = None
                if has_asset_id and pd.notna(row.get('Asset ID')):
                    asset = await db.assets.find_one({"asset_id": str(row['Asset ID'])}, {"_id": 0})
                elif has_serial_number and pd.notna(row.get('Asset Serial Number')):
                    asset = await db.assets.find_one({"serial_number": str(row['Asset Serial Number'])}, {"_id": 0})
                
                if not asset:
                    errors.append(f"Row {index + 2}: Asset not found")
                    continue
                
                # Check if asset is already assigned
                if asset["status"] == "Assigned":
                    errors.append(f"Row {index + 2}: Asset {asset['asset_id']} is already assigned")
                    continue
                
                # Generate or use provided Assignment ID
                assignment_id = None
                if 'Assignment ID' in df.columns and pd.notna(row.get('Assignment ID')):
                    assignment_id = str(row['Assignment ID'])
                    # Check if ID already exists
                    existing = await db.assignments.find_one({"assignment_id": assignment_id}, {"_id": 0})
                    if existing:
                        errors.append(f"Row {index + 2}: Assignment ID {assignment_id} already exists")
                        continue
                else:
                    count = await db.assignments.count_documents({})
                    assignment_id = f"ASG{str(count + 1).zfill(4)}"
                
                # Parse dates
                assigned_date = str(row['Assigned Date'])
                return_date = None
                if 'Return Date' in df.columns and pd.notna(row.get('Return Date')):
                    return_date = str(row['Return Date'])
                
                remarks = None
                if 'Remarks' in df.columns and pd.notna(row.get('Remarks')):
                    remarks = str(row['Remarks'])
                
                # Create assignment
                assignment_data = {
                    "assignment_id": assignment_id,
                    "employee_id": employee["employee_id"],
                    "employee_name": employee["full_name"],
                    "asset_id": asset["asset_id"],
                    "asset_name": asset["asset_name"],
                    "assigned_date": assigned_date,
                    "return_date": return_date,
                    "remarks": remarks
                }
                
                await db.assignments.insert_one(assignment_data)
                
                # Update asset status
                if return_date:
                    await db.assets.update_one({"asset_id": asset["asset_id"]}, {"$set": {"status": "Available"}})
                else:
                    await db.assets.update_one({"asset_id": asset["asset_id"]}, {"$set": {"status": "Assigned"}})
                
                imported_count += 1
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return {
            "message": f"Successfully imported {imported_count} asset assignments",
            "imported": imported_count,
            "errors": errors
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.get("/assignments/template")
async def download_assignments_template(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignments Template"
    
    headers = ["Employee ID", "Employee Email", "Asset ID", "Asset Serial Number", "Assigned Date", "Return Date", "Remarks"]
    ws.append(headers)
    
    # Add sample data with instructions
    ws.append(["EMP0001", "john@example.com", "AST0001", "SN123456", "2024-01-15", "", "New laptop for developer"])
    ws.append(["", "jane@example.com", "", "SN789012", "2024-01-16", "2024-01-20", "Returned in good condition"])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assignments_template.xlsx"}
    )

@api_router.get("/sim-connections")
async def get_sim_connections(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get all assignments with mobile assets
    assignments = await db.assignments.find({}, {"_id": 0}).to_list(1000)
    
    sim_connections = []
    for assignment in assignments:
        # Get asset to check category
        asset = await db.assets.find_one({"asset_id": assignment["asset_id"]}, {"_id": 0})
        
        if asset and asset.get("category", "").lower() == "mobile":
            # Only include if SIM details exist
            if assignment.get("sim_mobile_number") or assignment.get("sim_provider"):
                sim_connections.append({
                    "assignment_id": assignment.get("assignment_id"),
                    "sim_provider": assignment.get("sim_provider"),
                    "sim_mobile_number": assignment.get("sim_mobile_number"),
                    "sim_type": assignment.get("sim_type"),
                    "sim_ownership": assignment.get("sim_ownership"),
                    "sim_purpose": assignment.get("sim_purpose"),
                    "employee_name": assignment.get("employee_name"),
                    "asset_name": assignment.get("asset_name"),
                    "assigned_date": assignment.get("assigned_date"),
                    "return_date": assignment.get("return_date")
                })
    
    return sim_connections

@api_router.get("/sim-connections/export")
async def export_sim_connections(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get all SIM connections
    assignments = await db.assignments.find({}, {"_id": 0}).to_list(1000)
    
    sim_connections = []
    for assignment in assignments:
        asset = await db.assets.find_one({"asset_id": assignment["asset_id"]}, {"_id": 0})
        if asset and asset.get("category", "").lower() == "mobile":
            if assignment.get("sim_mobile_number") or assignment.get("sim_provider"):
                sim_connections.append(assignment)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "SIM Connections"
    
    headers = ["SIM Provider", "SIM Mobile Number", "SIM Type", "SIM Ownership", "SIM Purpose", 
               "Employee Name", "Asset Name", "Assigned Date", "Return Date"]
    ws.append(headers)
    
    for conn in sim_connections:
        ws.append([
            conn.get("sim_provider", ""),
            conn.get("sim_mobile_number", ""),
            conn.get("sim_type", ""),
            conn.get("sim_ownership", ""),
            conn.get("sim_purpose", ""),
            conn.get("employee_name", ""),
            conn.get("asset_name", ""),
            conn.get("assigned_date", ""),
            conn.get("return_date", "")
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sim_connections.xlsx"}
    )

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    total_assets = await db.assets.count_documents({})
    assigned_assets = await db.assets.count_documents({"status": "Assigned"})
    available_assets = await db.assets.count_documents({"status": "Available"})
    total_employees = await db.employees.count_documents({})
    
    return {
        "total_assets": total_assets,
        "assigned_assets": assigned_assets,
        "available_assets": available_assets,
        "total_employees": total_employees
    }

@api_router.get("/pending-returns", response_model=List[PendingReturn])
async def get_pending_returns(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    exit_employees = await db.employees.find({"status": "Exit"}, {"_id": 0}).to_list(1000)
    pending_returns = []
    
    for employee in exit_employees:
        unreturned_assets = await db.assignments.find(
            {
                "employee_id": employee["employee_id"],
                "return_date": None
            },
            {"_id": 0}
        ).to_list(1000)
        
        if unreturned_assets:
            pending_returns.append({
                "employee_id": employee["employee_id"],
                "employee_name": employee["full_name"],
                "email": employee["email"],
                "assets": unreturned_assets
            })
    
    return pending_returns

@api_router.get("/search/employees")
async def search_employees(q: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {
        "$or": [
            {"full_name": {"$regex": q, "$options": "i"}},
            {"employee_id": {"$regex": q, "$options": "i"}},
            {"department": {"$regex": q, "$options": "i"}}
        ]
    }
    employees = await db.employees.find(query, {"_id": 0}).to_list(100)
    
    for employee in employees:
        assignments = await db.assignments.find(
            {"employee_id": employee["employee_id"], "return_date": None},
            {"_id": 0}
        ).to_list(100)
        employee["assigned_assets"] = assignments
    
    return employees

@api_router.get("/assignments/export")
async def export_assignments(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    assignments = await db.assignments.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Assignments"
    
    headers = ["Assignment ID", "Employee ID", "Employee Name", "Asset ID", "Asset Name", 
               "Assigned Date", "Return Date", "Remarks",
               "SIM Provider", "SIM Mobile Number", "SIM Type", "SIM Ownership", "SIM Purpose"]
    ws.append(headers)
    
    for assignment in assignments:
        ws.append([
            assignment.get("assignment_id", ""),
            assignment.get("employee_id", ""),
            assignment.get("employee_name", ""),
            assignment.get("asset_id", ""),
            assignment.get("asset_name", ""),
            assignment.get("assigned_date", ""),
            assignment.get("return_date", ""),
            assignment.get("remarks", ""),
            assignment.get("sim_provider", ""),
            assignment.get("sim_mobile_number", ""),
            assignment.get("sim_type", ""),
            assignment.get("sim_ownership", ""),
            assignment.get("sim_purpose", "")
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=asset_assignments.xlsx"}
    )

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()